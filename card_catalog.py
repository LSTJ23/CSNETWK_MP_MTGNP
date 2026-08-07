import openpyxl

def load_card_catalog(excel_path="mtgnp_master_card_list.xlsx"):
    """
    Reads the Excel catalog and extracts:
    1. valid_card_ids: set of all 312 individual instance card_ids (e.g., 'mountain_001')
    2. base_cards: dict mapping base card IDs to properties (type, costs, power, toughness)
    """
    wb = openpyxl.load_workbook(excel_path)

    # 1. Load individual card instance IDs (data starts at row 3)
    sheet_instances = wb["Card Instances"]
    valid_card_ids = set()
    for row in sheet_instances.iter_rows(min_row=3, values_only=True):
        if row[0]:
            valid_card_ids.add(str(row[0]).strip())

    # 2. Load base card definitions (data starts at row 3)
    sheet_master = wb["Master Card List"]
    base_cards = {}
    for row in sheet_master.iter_rows(min_row=3, values_only=True):
        base_id = row[0]
        if base_id:
            base_cards[str(base_id).strip()] = {
                "name": row[1],
                "type": row[2],
                "subtype": row[3],
                "color": row[4],
                "cmc": row[5],
                "costs": {
                    "W": row[6], "U": row[7], "B": row[8],
                    "R": row[9], "G": row[10], "generic": row[11]
                },
                "power": row[12],
                "toughness": row[13],
                "copies": row[14],
                "effect": row[15]
            }

    return valid_card_ids, base_cards


if __name__ == "__main__":
    try:
        valid_ids, base_cards = load_card_catalog()
        print(f"[SUCCESS] Loaded {len(valid_ids)} valid instance card IDs.")
        print(f"[SUCCESS] Loaded {len(base_cards)} base card definitions.")
    except Exception as e:
        print(f"[ERROR] Could not load catalog: {e}")